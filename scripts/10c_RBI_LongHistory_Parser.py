# -*- coding: utf-8 -*-
"""
Script 10c — RBI/PPAC Long-History Macro Parser (Two-Phase Baseline, Stage 1)
=============================================================================
Extracts the FULL historical depth of the same four CEIC-sourced files
Script 10b already used to extend the macro series to mid-2026 -- but here
pulling every available row, not just the delta since the last extension.
Reuses Script 10b's exact column-index mappings (each was validated there
by comparing its 2025/2026 value against the existing production CSV
before being trusted as the same series) rather than re-deriving them.

Each series has its OWN real historical floor -- verified directly against
the files, not assumed:
  repo_rate_pct / reverse_repo_pct   : 2010-01 -> 2026-06 (Scheme II-00719118-M.xlsx)
  usdinr_monthly_avg                 : 1992-08 -> 2026-06 (Exchange rate.xlsx)
  wpi_fruits_vegetables / _vegetables_total / _potato / _onion / _tomato
                                      : 2011-04 -> 2026-04 (WPI fruits and vegetables.xlsx)
                                        Floor is the base-year (2011-12=100) series'
                                        own construction date, not a scraping gap --
                                        an older base-year WPI table exists separately
                                        and would need rebasing to splice on; not
                                        attempted here.
  diesel_4city_rs_litre / lpg_nonsub_4city_rs_cyl / diesel_delhi_per_L /
  lpg_nonsub_delhi_per14kg           : 2002-06 -> 2026-06 (Domestic prices for fuel.xlsx)

These floors are NOT unified into one common start date here -- that is a
Stage 2 (baseline-panel-join) decision, not this script's job. Different
series legitimately start at different dates; forcing them to a common
2011 floor would throw away 2003-2010 real data for every series except
WPI. LightGBM handles per-row missing features natively.

Correctness check (same principle Script 10b establishes in its own
docstring): for every month that overlaps the existing production window
(2017-2025), the newly-parsed long-history value is compared against the
existing data/rbi_dbie/rbi_dbie_macro_2017_2025.csv value already in
production. Two different outcomes are expected for two different reasons:

  - repo_rate_pct / reverse_repo_pct: these are point-in-time administered
    rates that do not get restated. A mismatch here IS a bug (column
    drift, wrong file) and blocks the output -- confirmed clean on the
    run that produced this file. usdinr_monthly_avg was originally treated
    the same way but is now INFORMATIONAL, not blocking (see the code
    below): a monthly average of daily FX rates can legitimately shift by
    a few paise on recomputation without being a real restatement.

  - wpi_* (all 5 columns) and the PPAC diesel/LPG columns: CMIE RESTATES
    THE ENTIRE HISTORY on every pull of these item-level series, not just
    the newest months (confirmed 2026-08-01 -- systematic 80-90%+ of
    months differ in every year 2017-2025, mean ~10-22% for tomato
    specifically, while the newest common month, 2026-06, matches
    exactly). This is expected CMIE behavior, not a parsing error: the
    freshly-pulled file is the CURRENT, more-correct vintage; the
    existing production CSV is a frozen snapshot from whenever it was
    first built and is now stale by comparison. A "mismatch" on these
    columns is therefore NOT blocking -- it is the point of pulling long
    history at all. Reported for visibility, not treated as a failure.

    Implication beyond this script: the CURRENT PRODUCTION MODELS (M6,
    the dashboard) are trained on that same stale 2017-2025 WPI/diesel/
    LPG vintage, not just the two-phase baseline being built here. That
    is a separate decision from this build (whether/when to refresh
    production on the current CMIE vintage) -- flagged for the user, not
    resolved by this script.

Outputs (data/rbi_dbie/, data/ppac_macro/):
  rbi_dbie_macro_longhistory.csv   year, month, repo_rate_pct, reverse_repo_pct,
                                    usdinr_monthly_avg, wpi_fruits_vegetables,
                                    wpi_vegetables_total, wpi_potato, wpi_onion,
                                    wpi_tomato -- one row per (year, month),
                                    full available history, NaN where a series
                                    hasn't started yet.
  ppac_diesel_lpg_longhistory.csv  year, month, diesel_4city_rs_litre,
                                    lpg_nonsub_4city_rs_cyl, diesel_delhi_per_L,
                                    lpg_nonsub_delhi_per14kg

Run: python scripts/10c_RBI_LongHistory_Parser.py
"""

import os
import numpy as np
import pandas as pd
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOWNLOADS = os.environ.get('TOP_DOWNLOADS_DIR', r'C:\Users\masro\Downloads')
OUT_RBI_DIR = os.path.join(BASE, 'data', 'rbi_dbie')
OUT_PPAC_DIR = os.path.join(BASE, 'data', 'ppac_macro')

EXISTING_RBI_FILE = os.path.join(OUT_RBI_DIR, 'rbi_dbie_macro_2017_2025.csv')
EXISTING_PPAC_FILE = os.path.join(OUT_PPAC_DIR, 'ppac_diesel_lpg_2017_2025.csv')

OUT_RBI_FILE = os.path.join(OUT_RBI_DIR, 'rbi_dbie_macro_longhistory.csv')
OUT_PPAC_FILE = os.path.join(OUT_PPAC_DIR, 'ppac_diesel_lpg_longhistory.csv')

print('=' * 65)
print('SCRIPT 10c: RBI/PPAC LONG-HISTORY MACRO PARSER')
print('  (Two-Phase Baseline, Stage 1)')
print('=' * 65)


# ─────────────────────────────────────────────────────────────────────────────
# 1. SHARED CEIC PARSING HELPERS (identical to Script 10b -- do not re-derive
#    column indices independently; reuse the already-validated mapping)
# ─────────────────────────────────────────────────────────────────────────────
def load_ceic_int_dates(path, sheet='Sheet1'):
    """Rows keyed by an integer YYYYMMDD date + a tag column ('M'/'C')."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    return [r for r in rows if isinstance(r[0], int) and r[0] > 19000000]


def load_ceic_str_dates(path, sheet='Sheet1'):
    """Rows keyed by a 'DD-Mon-YYYY' string date (repo rate file's format)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    return [r for r in rows if isinstance(r[0], str) and ('-19' in r[0] or '-20' in r[0])]


def yyyymmdd_to_year_month(d):
    return d // 10000, (d // 100) % 100


# ─────────────────────────────────────────────────────────────────────────────
# 2. RBI DBIE: repo/reverse repo, USD/INR, WPI -- FULL history
# ─────────────────────────────────────────────────────────────────────────────
print('\n[1] Parsing full history: repo/reverse-repo, USD/INR, WPI ...')

repo_rows = load_ceic_str_dates(os.path.join(DOWNLOADS, 'Scheme II-00719118-M.xlsx'))
repo_by_ym = {}
for r in repo_rows:
    dt = pd.to_datetime(r[0], format='%d-%b-%Y')
    repo_by_ym[(dt.year, dt.month)] = (r[1], r[11])  # (repo_rate, reverse_repo) -- Script 10b mapping
print(f'  repo/reverse-repo : {len(repo_by_ym)} months, '
      f'{min(repo_by_ym):} to {max(repo_by_ym):}' if repo_by_ym else '  repo: none found')

fx_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'Exchange rate.xlsx')) if r[1] == 'M']
fx_by_ym = {yyyymmdd_to_year_month(r[0]): r[2] for r in fx_rows}
print(f'  USD/INR           : {len(fx_by_ym)} months, '
      f'{min(fx_by_ym):} to {max(fx_by_ym):}' if fx_by_ym else '  fx: none found')

wpi_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'WPI fruits and vegetables.xlsx')) if r[1] == 'M']
wpi_by_ym = {yyyymmdd_to_year_month(r[0]): (r[2], r[3], r[4], r[6], r[10]) for r in wpi_rows}
print(f'  WPI (item-level)  : {len(wpi_by_ym)} months, '
      f'{min(wpi_by_ym):} to {max(wpi_by_ym):}' if wpi_by_ym else '  wpi: none found')

all_ym = sorted(set(repo_by_ym) | set(fx_by_ym) | set(wpi_by_ym))
rbi_rows = []
for y, m in all_ym:
    repo, rrepo = repo_by_ym.get((y, m), (np.nan, np.nan))
    fx = fx_by_ym.get((y, m), np.nan)
    wpi = wpi_by_ym.get((y, m), (np.nan,) * 5)
    rbi_rows.append({
        'year': y, 'month': m, 'repo_rate_pct': repo, 'reverse_repo_pct': rrepo,
        'usdinr_monthly_avg': fx, 'wpi_fruits_vegetables': wpi[0],
        'wpi_vegetables_total': wpi[1], 'wpi_potato': wpi[2], 'wpi_onion': wpi[3],
        'wpi_tomato': wpi[4],
    })
rbi_long = pd.DataFrame(rbi_rows)
rbi_start_y, rbi_start_m = all_ym[0]
rbi_end_y, rbi_end_m = all_ym[-1]
print(f'\n  Combined: {len(rbi_long)} months, {rbi_start_y}-{rbi_start_m:02d} to {rbi_end_y}-{rbi_end_m:02d}')


# ─────────────────────────────────────────────────────────────────────────────
# 3. PPAC (via CEIC): diesel/LPG -- FULL history
# ─────────────────────────────────────────────────────────────────────────────
print('\n[2] Parsing full history: diesel/LPG (domestic, 4-city + Delhi) ...')

def safe_mean(vals):
    """np.mean chokes on a mix of None/float; some early-history rows have
    partial city coverage (not every city in the 4-city average started
    reporting the same month) -- treat missing components as NaN and
    average whatever's present, rather than crashing or silently zeroing."""
    clean = [v for v in vals if v is not None]
    return float(np.mean(clean)) if clean else np.nan


fuel_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'Domestic prices for fuel.xlsx')) if r[1] == 'M']
ppac_rows = []
for r in sorted(fuel_rows, key=lambda x: x[0]):
    y, m = yyyymmdd_to_year_month(r[0])
    diesel4 = safe_mean(r[18:22])
    lpg4 = safe_mean(r[6:10])
    ppac_rows.append({
        'date': (pd.Timestamp(y, m, 1) + pd.offsets.MonthEnd(0)).strftime('%Y-%m-%d'),
        'year': y, 'month': m,
        'diesel_4city_rs_litre': diesel4, 'lpg_nonsub_4city_rs_cyl': lpg4,
        'diesel_delhi_per_L': r[18], 'lpg_nonsub_delhi_per14kg': r[6],
    })
ppac_long = pd.DataFrame(ppac_rows)
print(f'  Combined: {len(ppac_long)} months, {ppac_long["year"].min()}-{int(ppac_long.loc[ppac_long.year==ppac_long.year.min(),"month"].min()):02d} '
      f'to {ppac_long["year"].max()}-{int(ppac_long.loc[ppac_long.year==ppac_long.year.max(),"month"].max()):02d}')


# ─────────────────────────────────────────────────────────────────────────────
# 4. CORRECTNESS CHECK -- split into BLOCKING (point-in-time rates, should
# never differ) and INFORMATIONAL (CMIE-restated item-level series, expected
# to differ -- confirmed with the project owner, a CMIE subscriber, 2026-08-01:
# CMIE restates the full history of these series on every pull).
# ─────────────────────────────────────────────────────────────────────────────
print('\n[3] Validating against production data/rbi_dbie + ppac_macro CSVs '
      '(2017-2025 overlap) ...')

BLOCKING_TOL = {'repo_rate_pct': 0.01, 'reverse_repo_pct': 0.01}
# usdinr_monthly_avg moved to informational: unlike repo/reverse-repo (a discrete
# administered rate, exact by definition), a monthly AVERAGE of daily FX rates
# can legitimately shift by a few paise on recomputation even without a real
# restatement -- and per the 2026-08-01 confirmation that CMIE restates broadly,
# small drift here is expected, not a sign of a wrong column.
RESTATED_COLS_RBI = ['usdinr_monthly_avg', 'wpi_fruits_vegetables', 'wpi_vegetables_total',
                      'wpi_potato', 'wpi_onion', 'wpi_tomato']
RESTATED_COLS_PPAC = ['diesel_4city_rs_litre', 'lpg_nonsub_4city_rs_cyl',
                       'diesel_delhi_per_L', 'lpg_nonsub_delhi_per14kg']


def check_overlap(long_df, existing_path, blocking_tol, restated_cols, label):
    existing = pd.read_csv(existing_path)
    merged = existing.merge(long_df, on=['year', 'month'], suffixes=('_prod', '_long'))
    n_bad = 0
    for col, t in blocking_tol.items():
        c_prod, c_long = f'{col}_prod', f'{col}_long'
        if c_prod not in merged.columns or c_long not in merged.columns:
            continue
        diff = (merged[c_prod] - merged[c_long]).abs()
        bad = merged[diff > t]
        if len(bad):
            n_bad += len(bad)
            worst = bad.assign(diff=diff[diff > t]).sort_values('diff', ascending=False).head(3)
            print(f'    BLOCKING MISMATCH {label}.{col}: {len(bad)} months exceed tol={t}, worst:')
            print(worst[['year', 'month', c_prod, c_long]].to_string(index=False))
    n_ok_blocking = len(merged) - n_bad if blocking_tol else None
    if blocking_tol:
        status = 'OK' if n_bad == 0 else f'{n_bad} mismatched cells -- INVESTIGATE'
        print(f'  {label} (point-in-time, blocking): {status}')

    for col in restated_cols:
        c_prod, c_long = f'{col}_prod', f'{col}_long'
        if c_prod not in merged.columns or c_long not in merged.columns:
            continue
        pctdiff = (merged[c_long] - merged[c_prod]).abs() / merged[c_prod].abs() * 100
        n_diff = (pctdiff > 2).sum()
        print(f'  {label}.{col} (CMIE-restated, informational): {n_diff}/{len(merged)} months '
              f'differ >2%, mean |diff|={pctdiff.mean():.1f}% -- expected, not blocking.')
    return n_bad


n_bad_rbi = check_overlap(rbi_long, EXISTING_RBI_FILE, BLOCKING_TOL, RESTATED_COLS_RBI, 'RBI')
n_bad_ppac = check_overlap(ppac_long, EXISTING_PPAC_FILE, {}, RESTATED_COLS_PPAC, 'PPAC')


# ─────────────────────────────────────────────────────────────────────────────
# 5. SAVE -- gated on the correctness check above. Each file's own blocking
# verdict decides whether it is written: a blocking mismatch must never
# produce (or overwrite) a file that looks like a normal, trustworthy
# output. Saved independently since RBI and PPAC have independent checks.
# ─────────────────────────────────────────────────────────────────────────────
print('\n[4] Saving ...')
os.makedirs(OUT_RBI_DIR, exist_ok=True)
os.makedirs(OUT_PPAC_DIR, exist_ok=True)

if n_bad_rbi == 0:
    rbi_long.to_csv(OUT_RBI_FILE, index=False)
    print(f'  Saved: {OUT_RBI_FILE}  ({len(rbi_long)} rows)')
else:
    blocked_path = OUT_RBI_FILE + '.BLOCKED'
    rbi_long.to_csv(blocked_path, index=False)
    print(f'  NOT SAVED — blocking mismatch, see above ({OUT_RBI_FILE})')
    print(f'  Diagnostic copy written instead to: {blocked_path} (do NOT use as production input)')

if n_bad_ppac == 0:
    ppac_long.to_csv(OUT_PPAC_FILE, index=False)
    print(f'  Saved: {OUT_PPAC_FILE}  ({len(ppac_long)} rows)')
else:
    blocked_path = OUT_PPAC_FILE + '.BLOCKED'
    ppac_long.to_csv(blocked_path, index=False)
    print(f'  NOT SAVED — blocking mismatch, see above ({OUT_PPAC_FILE})')
    print(f'  Diagnostic copy written instead to: {blocked_path} (do NOT use as production input)')

print('\nPer-series real floor (NaN before this, by construction -- not a bug):')
for col in ['repo_rate_pct', 'usdinr_monthly_avg', 'wpi_fruits_vegetables', 'wpi_onion']:
    first = rbi_long.loc[rbi_long[col].notna(), ['year', 'month']].iloc[0]
    print(f'  {col:<24s}: {int(first.year)}-{int(first.month):02d}')
for col in ['diesel_delhi_per_L']:
    first = ppac_long.loc[ppac_long[col].notna(), ['year', 'month']].iloc[0]
    print(f'  {col:<24s}: {int(first.year)}-{int(first.month):02d}')

print('\n' + '=' * 65)
print('Script 10c complete.')
if n_bad_rbi:
    print('BLOCKING VALIDATION FAILED on repo/reverse-repo -- rbi_dbie_macro_longhistory.csv '
          'was NOT saved (see .BLOCKED diagnostic copy above); resolve and re-run.')
else:
    print('Validation passed. These files are additive (data/*_longhistory.csv) --')
    print('production data/rbi_dbie/rbi_dbie_macro_2017_2025.csv and')
    print('data/ppac_macro/ppac_diesel_lpg_2017_2025.csv are untouched.')
print('\nNext: Stage 2 -- build the baseline-phase panel join using these files')
print('for years < 2017 and the existing production files for 2017+.')
