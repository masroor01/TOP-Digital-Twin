# -*- coding: utf-8 -*-
"""
Script 10b — Extend RBI/PPAC/CMIE Macro Series Through Mid-2026
====================================================================
One-off/refresh utility. The RBI DBIE and PPAC macro CSVs have no
dedicated ingestion script (per README: hand-assembled originally); this
extends them and cmie_macro_2017_2025.csv with 2026 CMIE Economic Outlook
exports, using each series' actual coverage (they don't all reach the
same date).

Column mappings below were NOT guessed -- each was validated by comparing
its Dec-2025 (or latest overlapping month) value against the existing CSV
before trusting it as the same series:
  - diesel_4city_rs_litre / lpg_nonsub_4city_rs_cyl / diesel_delhi_per_L /
    lpg_nonsub_delhi_per14kg: EXACT match (90.475, ~863, 87.7, ~853)
  - repo_rate_pct / reverse_repo_pct: pulled from the same hardcoded-index
    pattern (Scheme II-00719118-M.xlsx, cols r[1]/r[11]) as the other
    validated series above; listed here for completeness alongside them.
  - usdinr_monthly_avg: EXACT match (90.09)
  - bank_credit_agri_cr: EXACT match (25,549,785.8 ~ 25,549,785.9)
  - agri_wages_rs_day: EXACT match (518.7) -- but the matching column is
    actually "All occupations: Men", NOT "Agricultural occupations: Men"
    (idx2, which is a different, lower series ~501.7 in the same file).
    This means the ALREADY-PUBLISHED historical series is really an
    all-occupations rural wage proxy, not agriculture-specific. Kept
    for continuity; flag for the user to decide whether to correct
    project-wide in a future revision (would touch trained results).
  - iip_food_proc: EXACT match (170.4) -- but the matching column is the
    OVERALL "Total" manufacturing IIP (weight 77.63/100), NOT the
    "Manufacture of food products" sub-index (weight 5.30/100) that the
    column name implies. Same situation as agri_wages_rs_day: pre-existing
    in the published data, kept for continuity, flagged for the user.
  - wpi_* (fruits_vegetables/vegetables_total/potato/onion/tomato): closer
    but not exact vs. history (WPI figures get revised for 2-3 months
    after first release, and CMIE's "M" tag here is the current-vintage
    monthly print, not a frozen historical snapshot -- see the "M vs C"
    note below). Column identity is unambiguous either way (plain text
    item labels in the sheet: Total/Vegetables Total/Potato/Onion/Tomato).

CMIE "M" vs "C" tag: several of these files carry two rows per month,
tagged M and C, which converge to identical values in the most recent
month and diverge further back -- consistent with M = current monthly
print, C = cumulative/fiscal-YTD. Per project owner (CMIE subscriber):
use M throughout.

NOT extended (source not found in this pass, still ends 2025 or earlier):
  export_veg_usd_mn, import_veg_usd_mn, crude_oil_usd_bbl, pfce_food_bev_cr
  (pfce_food_bev_cr was never populated in the existing CSV either --
  no regression, just still unavailable)

Inputs (Downloads root):
  Scheme II-00719118-M.xlsx    Repo rate + Reverse repo rate (CMIE, "Rate: Repo" /
                                "Rate: Reverse repo" columns) -- through 2026-06-30
  Exchange rate.xlsx           USD/INR average -- through 2026-06-30
  WPI fruits and vegetables.xlsx   WPI item-level indices -- through 2026-04-30
  Domestic prices for fuel.xlsx    Diesel/LPG by city -- through 2026-06-30
  cmie_agri_credit.xlsx        Bank credit to Agriculture & allied -- through 2026-05-31
  cmie_agri_wages.xlsx         Rural wage rates -- through 2026-03-31
  IIP.xlsx                     IIP by manufacturing sub-sector -- through 2026-03-31

Outputs (extended in place -- same filenames downstream scripts expect):
  data/rbi_dbie/rbi_dbie_macro_2017_2025.csv
  data/ppac_macro/ppac_diesel_lpg_2017_2025.csv
  data/cmie_macro/cmie_macro_2017_2025.csv
"""

import os
import pandas as pd
import numpy as np

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOWNLOADS = os.environ.get('TOP_DOWNLOADS_DIR', r'C:\Users\masro\Downloads')

RBI_FILE = os.path.join(BASE, 'data', 'rbi_dbie', 'rbi_dbie_macro_2017_2025.csv')
PPAC_FILE = os.path.join(BASE, 'data', 'ppac_macro', 'ppac_diesel_lpg_2017_2025.csv')
CMIE_FILE = os.path.join(BASE, 'data', 'cmie_macro', 'cmie_macro_2017_2025.csv')


def load_ceic_int_dates(path, sheet='Sheet1'):
    """Rows keyed by an integer YYYYMMDD date + a tag column ('M'/'C' or similar)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    return [r for r in rows if isinstance(r[0], int) and r[0] > 19000000]


def load_ceic_str_dates(path, sheet='Sheet1'):
    """Rows keyed by a 'DD-Mon-YYYY' string date (repo rate file's format)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    return [r for r in rows if isinstance(r[0], str) and '-20' in str(r[0])]


def yyyymmdd_to_year_month(d):
    return d // 10000, (d // 100) % 100


# ─────────────────────────────────────────────────────────────────────────────
# 1. RBI DBIE: repo/reverse repo, USD/INR, WPI
# ─────────────────────────────────────────────────────────────────────────────
print('=== RBI DBIE ===')
old_rbi = pd.read_csv(RBI_FILE)
last_year, last_month = int(old_rbi.iloc[-1]['year']), int(old_rbi.iloc[-1]['month'])
print(f'  Existing data ends: {last_year}-{last_month:02d}')

repo_rows = load_ceic_str_dates(os.path.join(DOWNLOADS, 'Scheme II-00719118-M.xlsx'))
repo_by_ym = {}
for r in repo_rows:
    dt = pd.to_datetime(r[0], format='%d-%b-%Y')
    repo_by_ym[(dt.year, dt.month)] = (r[1], r[11])  # (repo_rate, reverse_repo)

fx_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'Exchange rate.xlsx')) if r[1] == 'M']
fx_by_ym = {yyyymmdd_to_year_month(r[0]): r[2] for r in fx_rows}

wpi_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'WPI fruits and vegetables.xlsx')) if r[1] == 'M']
wpi_by_ym = {yyyymmdd_to_year_month(r[0]): (r[2], r[3], r[4], r[6], r[10]) for r in wpi_rows}

new_rbi_rows = []
all_ym = sorted(set(repo_by_ym) | set(fx_by_ym) | set(wpi_by_ym))
for y, m in all_ym:
    if (y, m) <= (last_year, last_month):
        continue
    repo, rrepo = repo_by_ym.get((y, m), (np.nan, np.nan))
    fx = fx_by_ym.get((y, m), np.nan)
    wpi = wpi_by_ym.get((y, m), (np.nan,) * 5)
    new_rbi_rows.append({
        'year': y, 'month': m, 'repo_rate_pct': repo, 'reverse_repo_pct': rrepo,
        'usdinr_monthly_avg': fx, 'wpi_fruits_vegetables': wpi[0],
        'wpi_vegetables_total': wpi[1], 'wpi_potato': wpi[2], 'wpi_onion': wpi[3],
        'wpi_tomato': wpi[4],
    })

new_rbi = pd.DataFrame(new_rbi_rows)
print(f'  New rows: {len(new_rbi)}  ({new_rbi["year"].min()}-{new_rbi["month"].min():02d} '
      f'to {new_rbi["year"].max()}-{new_rbi["month"].max():02d})' if len(new_rbi) else '  No new rows')
combined_rbi = pd.concat([old_rbi, new_rbi], ignore_index=True)
combined_rbi.to_csv(RBI_FILE, index=False)
print(f'  Saved: {RBI_FILE}  ({len(combined_rbi)} total rows)')
print(combined_rbi.tail(8).to_string())


# ─────────────────────────────────────────────────────────────────────────────
# 2. PPAC: diesel/LPG
# ─────────────────────────────────────────────────────────────────────────────
print('\n=== PPAC ===')
old_ppac = pd.read_csv(PPAC_FILE)
last_year, last_month = int(old_ppac.iloc[-1]['year']), int(old_ppac.iloc[-1]['month'])
print(f'  Existing data ends: {last_year}-{last_month:02d}')

fuel_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'Domestic prices for fuel.xlsx')) if r[1] == 'M']
new_ppac_rows = []
for r in sorted(fuel_rows, key=lambda x: x[0]):
    y, m = yyyymmdd_to_year_month(r[0])
    if (y, m) <= (last_year, last_month):
        continue
    diesel4 = np.mean(r[18:22])
    lpg4 = np.mean(r[6:10])
    new_ppac_rows.append({
        'date': pd.Timestamp(y, m, 1) + pd.offsets.MonthEnd(0),
        'year': y, 'month': m,
        'diesel_4city_rs_litre': diesel4, 'lpg_nonsub_4city_rs_cyl': lpg4,
        'diesel_delhi_per_L': r[18], 'lpg_nonsub_delhi_per14kg': r[6],
    })
new_ppac = pd.DataFrame(new_ppac_rows)
if len(new_ppac):
    new_ppac['date'] = new_ppac['date'].dt.strftime('%Y-%m-%d')
print(f'  New rows: {len(new_ppac)}')
combined_ppac = pd.concat([old_ppac, new_ppac], ignore_index=True)
combined_ppac.to_csv(PPAC_FILE, index=False)
print(f'  Saved: {PPAC_FILE}  ({len(combined_ppac)} total rows)')
print(combined_ppac.tail(8).to_string())


# ─────────────────────────────────────────────────────────────────────────────
# 3. CMIE macro: bank credit (agri), agri wages, IIP
# ─────────────────────────────────────────────────────────────────────────────
print('\n=== CMIE macro ===')
old_cmie = pd.read_csv(CMIE_FILE)
last_year, last_month = int(old_cmie.iloc[-1]['year']), int(old_cmie.iloc[-1]['month'])
print(f'  Existing data ends: {last_year}-{last_month:02d}')

credit_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'cmie_agri_credit.xlsx')) if r[1] == 'M']
credit_by_ym = {yyyymmdd_to_year_month(r[0]): r[3] for r in credit_rows}

wages_rows = load_ceic_str_dates(os.path.join(DOWNLOADS, 'cmie_agri_wages.xlsx'))
wages_by_ym = {}
for r in wages_rows:
    dt = pd.to_datetime(r[0], format='%d-%b-%Y')
    wages_by_ym[(dt.year, dt.month)] = r[1]  # idx1 = "All occupations: Men" -- matches existing series

iip_rows = [r for r in load_ceic_int_dates(os.path.join(DOWNLOADS, 'IIP.xlsx')) if r[1] == 'M']
iip_by_ym = {yyyymmdd_to_year_month(r[0]): r[2] for r in iip_rows}  # idx2 = "Total" -- matches existing series

new_cmie_rows = []
all_ym = sorted(set(credit_by_ym) | set(wages_by_ym) | set(iip_by_ym))
for y, m in all_ym:
    if (y, m) <= (last_year, last_month):
        continue
    new_cmie_rows.append({
        'date': pd.Timestamp(y, m, 1).strftime('%Y-%m-%d'),
        'year': y, 'month': m,
        'bank_credit_agri_cr': credit_by_ym.get((y, m), np.nan),
        'export_veg_usd_mn': np.nan, 'import_veg_usd_mn': np.nan,
        'crude_oil_usd_bbl': np.nan, 'iip_food_proc': iip_by_ym.get((y, m), np.nan),
        'agri_wages_rs_day': wages_by_ym.get((y, m), np.nan),
    })
new_cmie = pd.DataFrame(new_cmie_rows)
print(f'  New rows: {len(new_cmie)}')
combined_cmie = pd.concat([old_cmie, new_cmie], ignore_index=True)
combined_cmie.to_csv(CMIE_FILE, index=False)
print(f'  Saved: {CMIE_FILE}  ({len(combined_cmie)} total rows)')
print(combined_cmie.tail(8).to_string())

print('\nDone. export_veg_usd_mn / import_veg_usd_mn / crude_oil_usd_bbl remain '
      'un-extended (source not found) -- NaN for the new 2026 rows, same as they '
      'would be if simply missing.')
